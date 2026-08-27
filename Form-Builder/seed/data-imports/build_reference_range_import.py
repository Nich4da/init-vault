#!/usr/bin/env python3
"""Build the guarded initCraft Excel import for the CHE reference-range form."""

from __future__ import annotations

import json
import warnings
from collections import Counter
from pathlib import Path
from typing import Any

from numbers_parser import Document
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


WORKSPACE = Path("/Users/nichada/Documents/codex-backup")
SOURCE_NUMBERS = Path("/Users/nichada/Documents/Reference range Chem 13 July 26.numbers")
LAB_TEST_EXPORT = Path(
    "/Users/nichada/.mongodb/mongodb-mcp/exports/"
    "6a8eb53fbd2698db74a05cc2/6a8ec412bd2698db74a05cc4.json"
)
UNIT_EXPORT = Path(
    "/Users/nichada/.mongodb/mongodb-mcp/exports/"
    "6a8eb53fbd2698db74a05cc2/6a8ec41ebd2698db74a05cc5.json"
)
OUTPUT = WORKSPACE / "Reference_Range_Chem_initCraft_Import_2026-08-26.xlsx"


SERVICE_TYPE = {
    "value": "6a58f246d448dfc9d33e2bed",
    "label": "Lab Test",
    "code": "lab",
    "name": "Lab Test",
}
SECTION = {
    "value": "6a58f7f4d448dfc9d33e2bf3",
    "label": "Biochemistry",
    "modality_type": "",
    "name": "Biochemistry",
    "st_id": {
        "value": "6a58f246d448dfc9d33e2bed",
        "label": "Lab Test",
    },
    "code": "BC",
}

PHYSIOLOGICAL_TYPE = {"ALL": "All", "M": "M", "F": "F"}
INITIAL_AGE_UNIT = {
    "DAYS": "start_day",
    "MONTHS": "start_months",
    "YEARS": "start_year",
}
FINAL_AGE_UNIT = {
    "DAYS": "final_day",
    "MONTHS": "final_months",
    "YEARS": "final_year",
}
NUMERIC_OPERATOR = {
    "Inclusive limits": "Inclusive",
    "Exclusive limits": "Exclusive",
    "<": "lessthan",
    "<=": "lessthan-equal",
    ">": "greaterthan",
    ">=": "greaterthan-equal",
}
RANGE_TYPE = {"NormalNumeric": "normal", "CriticalNumeric": "critical"}

IMPORT_HEADERS = [
    "service_type",
    "section",
    "source_test_code",
    "source_test_reference",
    "lab_item",
    "revision_no",
    "is_active",
    "reference_ranges",
]

RAW_HEADERS = [
    "source_row",
    "source_test_code",
    "source_test_reference",
    "priority",
    "physiological_type_source",
    "physiological_type_import",
    "method",
    "initial_age",
    "initial_age_unit_source",
    "initial_age_unit_import",
    "final_age",
    "final_age_unit_source",
    "final_age_unit_import",
    "range_unit_source",
    "range_unit_master_id_exact",
    "range_unit_master_symbol_exact",
    "numeric_operator_source",
    "numeric_operator_import",
    "range_type_source",
    "range_type_import",
    "first_value",
    "second_value",
    "max_result_status_id",
    "min_result_status_id",
    "blocking_reasons",
]


def compact_json(value: Any) -> str:
    return json.dumps(value, ensure_ascii=False, separators=(",", ":"))


def clean_text(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def number_text(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, str):
        value = value.strip()
        return value or None
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return format(value, ".15g")
    return str(value)


def integer_value(value: Any) -> int | None:
    if value is None:
        return None
    if isinstance(value, str) and not value.strip():
        return None
    return int(float(value))


def load_source() -> list[dict[str, Any]]:
    warnings.filterwarnings("ignore", message="unsupported version")
    table = Document(SOURCE_NUMBERS).sheets[0].tables[0]
    rows = [[table.cell(row, col).value for col in range(table.num_cols)] for row in range(1, table.num_rows)]

    groups: list[dict[str, Any]] = []
    current: dict[str, Any] | None = None
    for source_row, row in enumerate(rows, start=2):
        if row[0] not in (None, ""):
            code = number_text(row[0])
            current = {
                "code": code,
                "reference": clean_text(row[1]),
                "method": clean_text(row[4]),
                "rules": [],
            }
            groups.append(current)
        if current is None:
            raise ValueError(f"Source row {source_row} appears before the first Test code")
        current["rules"].append(
            {
                "source_row": source_row,
                "priority": integer_value(row[2]),
                "physiological_type_source": clean_text(row[3]),
                "method": clean_text(row[4]),
                "initial_age": integer_value(row[5]),
                "initial_age_unit_source": clean_text(row[6]),
                "final_age": integer_value(row[7]),
                "final_age_unit_source": clean_text(row[8]),
                "range_unit_source": clean_text(row[11]),
                "numeric_operator_source": clean_text(row[13]),
                "range_type_source": clean_text(row[14]),
                "first_value": number_text(row[15]),
                "second_value": number_text(row[16]),
                "max_result_status_id": clean_text(row[17]),
                "min_result_status_id": clean_text(row[18]),
            }
        )
    return groups


def load_masters() -> tuple[dict[str, dict[str, Any]], dict[str, dict[str, Any]]]:
    lab_tests = json.loads(LAB_TEST_EXPORT.read_text(encoding="utf-8"))
    units = json.loads(UNIT_EXPORT.read_text(encoding="utf-8"))
    by_code = {str(row["code"]): row for row in lab_tests}
    by_symbol = {str(row["unit_symbol"]): row for row in units}
    return by_code, by_symbol


def oid(row: dict[str, Any]) -> str:
    value = row.get("_id")
    if isinstance(value, dict):
        return str(value.get("$oid") or "")
    return str(value or "")


def rule_issues(rule: dict[str, Any], unit_by_symbol: dict[str, dict[str, Any]]) -> list[str]:
    issues: list[str] = []
    if rule["physiological_type_source"] not in PHYSIOLOGICAL_TYPE:
        issues.append("PHYSIOLOGY_UNMAPPED")
    if rule["initial_age_unit_source"] not in INITIAL_AGE_UNIT:
        issues.append("INITIAL_AGE_UNIT_UNMAPPED")
    if rule["final_age_unit_source"] not in FINAL_AGE_UNIT:
        issues.append("FINAL_AGE_UNIT_UNMAPPED")
    if rule["numeric_operator_source"] not in NUMERIC_OPERATOR:
        issues.append("OPERATOR_UNMAPPED")
    if rule["range_type_source"] not in RANGE_TYPE:
        issues.append("RANGE_TYPE_UNMAPPED")
    source_unit = rule["range_unit_source"]
    if source_unit is not None and source_unit not in unit_by_symbol:
        issues.append("UNIT_NOT_EXACT")
    if rule["first_value"] is None:
        issues.append("FIRST_VALUE_MISSING")
    return issues


def lab_item_value(row: dict[str, Any]) -> dict[str, Any]:
    code = str(row["code"])
    label_parts = [code, clean_text(row.get("name")), clean_text(row.get("flag"))]
    return {
        "value": code,
        "label": " ".join(part for part in label_parts if part),
        "code": code,
        "name": row.get("name"),
        "short_name": row.get("short_name"),
        "flag": row.get("flag"),
        "is_show_chart": row.get("is_show_chart"),
        "section": row.get("section"),
        "specimen": row.get("specimen"),
        "seq": row.get("seq"),
    }


def import_rule(rule: dict[str, Any], unit_by_symbol: dict[str, dict[str, Any]]) -> dict[str, Any]:
    source_unit = rule["range_unit_source"]
    unit = unit_by_symbol.get(source_unit) if source_unit else None
    return {
        "priority": rule["priority"],
        "physiological_type": PHYSIOLOGICAL_TYPE[rule["physiological_type_source"]],
        "range_unit_id": {"value": oid(unit), "label": source_unit} if unit else None,
        "numeric_range_operator": NUMERIC_OPERATOR[rule["numeric_operator_source"]],
        "initial_age": rule["initial_age"],
        "initial_age_unit": INITIAL_AGE_UNIT[rule["initial_age_unit_source"]],
        "final_age": rule["final_age"],
        "final_age_unit": FINAL_AGE_UNIT[rule["final_age_unit_source"]],
        "range_type": RANGE_TYPE[rule["range_type_source"]],
        "first_value": rule["first_value"],
        "second_value": rule["second_value"],
        "min_result_status_id": rule["min_result_status_id"],
        "max_result_status_id": rule["max_result_status_id"],
    }


def set_sheet_style(ws, widths: dict[str, float] | None = None) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    if widths:
        for column, width in widths.items():
            ws.column_dimensions[column].width = width
    else:
        for index in range(1, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(index)].width = 20


def write_readme(wb: Workbook, ready_count: int, ready_rules: int, blocked_count: int, blocked_rules: int) -> None:
    ws = wb.create_sheet("README")
    rows = [
        ("หัวข้อ", "รายละเอียด"),
        ("ไฟล์ปลายทาง", "initCraft form #6a8dd522f851000f28e502c3 (referance range)"),
        ("ชีตที่ import", "IMPORT_READY_78 — ต้องเลือก/อ่านชีตแรกเท่านั้น"),
        ("จำนวนพร้อม import", f"{ready_count} parent records / {ready_rules} reference-range rules"),
        ("จำนวนพักไว้ตรวจ", f"{blocked_count} test groups / {blocked_rules} rules; ดู REVIEW_54_TESTS และ REVIEW_329_RULES"),
        ("รูปแบบข้อมูล", "หนึ่งแถวต่อหนึ่ง Test; reference_ranges เป็น JSON array สำหรับ Sub Form"),
        ("การ map", "map เฉพาะ 8 field ตามชื่อ header ใน IMPORT_READY_78; ไม่ map field แบบ reference_ranges.* ซ้ำ"),
        ("Select By Form", "service_type, section, lab_item และ range_unit_id เก็บ JSON {value,label,...}; ObjectId มาจาก live master แบบ read-only"),
        ("หน่วย", "นำเข้าเฉพาะ unit_symbol ที่ตรงกับ Lab_Unit_Master แบบ exact; ไม่แปลง alias โดยเดา"),
        ("ข้อจำกัด", "Static checks ผ่าน แต่ยังไม่ได้กด Import/เปิด record ใน initCraft runtime"),
        ("หลัง import", "สุ่มเปิดอย่างน้อย 1 record หลายช่วงอายุ ตรวจ 3 Tabs, Sub Form, เพศ, อายุ, operator, unit และค่าขอบเขต"),
    ]
    for row in rows:
        ws.append(row)
    set_sheet_style(ws, {"A": 24, "B": 120})


def build() -> dict[str, int]:
    groups = load_source()
    lab_by_code, unit_by_symbol = load_masters()

    ready: list[dict[str, Any]] = []
    blocked: list[dict[str, Any]] = []
    for group in groups:
        issues = []
        if group["code"] not in lab_by_code:
            issues.append("LAB_ITEM_NOT_FOUND")
        for rule in group["rules"]:
            issues.extend(rule_issues(rule, unit_by_symbol))
        group["issues"] = list(dict.fromkeys(issues))
        (blocked if group["issues"] else ready).append(group)

    wb = Workbook()
    import_ws = wb.active
    import_ws.title = "IMPORT_READY_78"
    import_ws.append(IMPORT_HEADERS)
    for group in ready:
        lab = lab_by_code[group["code"]]
        ranges = [import_rule(rule, unit_by_symbol) for rule in group["rules"]]
        values = [
            compact_json(SERVICE_TYPE),
            compact_json(SECTION),
            group["code"],
            group["reference"],
            compact_json(lab_item_value(lab)),
            1,
            True,
            compact_json(ranges),
        ]
        if any(isinstance(value, str) and len(value) > 32767 for value in values):
            raise ValueError(f"Excel cell length limit exceeded for Test {group['code']}")
        import_ws.append(values)
    set_sheet_style(
        import_ws,
        {"A": 45, "B": 55, "C": 18, "D": 28, "E": 95, "F": 12, "G": 12, "H": 120},
    )

    review_test_ws = wb.create_sheet("REVIEW_54_TESTS")
    review_test_ws.append(["source_test_code", "source_test_reference", "rule_count", "blocking_reasons", "action"])
    actions = {
        "LAB_ITEM_NOT_FOUND": "เพิ่ม/ยืนยัน Lab Test master code ก่อน import",
        "UNIT_NOT_EXACT": "ยืนยัน mapping unit alias หรือเพิ่ม unit_symbol ใน Lab_Unit_Master",
        "INITIAL_AGE_UNIT_UNMAPPED": "เพิ่ม option/ยืนยันความหมายของช่วงอายุเริ่มต้น",
        "FINAL_AGE_UNIT_UNMAPPED": "เพิ่ม option/ยืนยันความหมายของช่วงอายุสิ้นสุด",
        "OPERATOR_UNMAPPED": "ยืนยันความหมาย source operator '-'",
        "RANGE_TYPE_UNMAPPED": "เพิ่ม/ยืนยัน option CustomNumeric",
        "PHYSIOLOGY_UNMAPPED": "ยืนยัน physiological type",
        "FIRST_VALUE_MISSING": "เติมหรือยืนยัน First Value",
    }
    for group in blocked:
        review_test_ws.append(
            [
                group["code"],
                group["reference"],
                len(group["rules"]),
                " | ".join(group["issues"]),
                " ; ".join(actions[issue] for issue in group["issues"]),
            ]
        )
    set_sheet_style(review_test_ws, {"A": 20, "B": 28, "C": 12, "D": 70, "E": 110})

    review_rule_ws = wb.create_sheet("REVIEW_329_RULES")
    review_rule_ws.append(RAW_HEADERS)
    for group in blocked:
        for rule in group["rules"]:
            per_rule = rule_issues(rule, unit_by_symbol)
            if "LAB_ITEM_NOT_FOUND" in group["issues"]:
                per_rule.insert(0, "LAB_ITEM_NOT_FOUND")
            source_unit = rule["range_unit_source"]
            unit = unit_by_symbol.get(source_unit) if source_unit else None
            review_rule_ws.append(
                [
                    rule["source_row"],
                    group["code"],
                    group["reference"],
                    rule["priority"],
                    rule["physiological_type_source"],
                    PHYSIOLOGICAL_TYPE.get(rule["physiological_type_source"]),
                    rule["method"],
                    rule["initial_age"],
                    rule["initial_age_unit_source"],
                    INITIAL_AGE_UNIT.get(rule["initial_age_unit_source"]),
                    rule["final_age"],
                    rule["final_age_unit_source"],
                    FINAL_AGE_UNIT.get(rule["final_age_unit_source"]),
                    source_unit,
                    oid(unit) if unit else None,
                    unit.get("unit_symbol") if unit else None,
                    rule["numeric_operator_source"],
                    NUMERIC_OPERATOR.get(rule["numeric_operator_source"]),
                    rule["range_type_source"],
                    RANGE_TYPE.get(rule["range_type_source"]),
                    rule["first_value"],
                    rule["second_value"],
                    rule["max_result_status_id"],
                    rule["min_result_status_id"],
                    " | ".join(dict.fromkeys(per_rule)),
                ]
            )
    set_sheet_style(review_rule_ws)

    lab_map_ws = wb.create_sheet("LAB_TEST_MAP_132")
    lab_map_ws.append(
        ["source_test_code", "source_test_reference", "match_status", "lab_test_id", "lab_test_name", "short_name", "flag", "section_code"]
    )
    for group in groups:
        lab = lab_by_code.get(group["code"])
        lab_map_ws.append(
            [
                group["code"],
                group["reference"],
                "EXACT_CODE" if lab else "NOT_FOUND",
                oid(lab) if lab else None,
                lab.get("name") if lab else None,
                lab.get("short_name") if lab else None,
                lab.get("flag") if lab else None,
                (lab.get("section") or {}).get("code") if lab else None,
            ]
        )
    set_sheet_style(lab_map_ws, {"A": 20, "B": 28, "C": 16, "D": 28, "E": 42, "F": 18, "G": 12, "H": 14})

    unit_map_ws = wb.create_sheet("UNIT_MAP_21")
    unit_map_ws.append(["source_unit", "source_rule_count", "match_status", "unit_master_id", "unit_code", "unit_symbol", "note"])
    unit_counts = Counter(rule["range_unit_source"] for group in groups for rule in group["rules"])
    for source_unit in sorted(unit_counts, key=lambda value: (value is None, value or "")):
        unit = unit_by_symbol.get(source_unit) if source_unit else None
        unit_map_ws.append(
            [
                source_unit,
                unit_counts[source_unit],
                "EXACT_SYMBOL" if unit else ("SOURCE_BLANK" if source_unit is None else "REVIEW_REQUIRED"),
                oid(unit) if unit else None,
                unit.get("unit_code") if unit else None,
                unit.get("unit_symbol") if unit else None,
                "ไม่เดา alias" if source_unit and not unit else None,
            ]
        )
    set_sheet_style(unit_map_ws, {"A": 24, "B": 18, "C": 22, "D": 28, "E": 24, "F": 24, "G": 32})

    write_readme(
        wb,
        ready_count=len(ready),
        ready_rules=sum(len(group["rules"]) for group in ready),
        blocked_count=len(blocked),
        blocked_rules=sum(len(group["rules"]) for group in blocked),
    )

    wb.save(OUTPUT)
    return {
        "source_tests": len(groups),
        "source_rules": sum(len(group["rules"]) for group in groups),
        "ready_tests": len(ready),
        "ready_rules": sum(len(group["rules"]) for group in ready),
        "blocked_tests": len(blocked),
        "blocked_rules": sum(len(group["rules"]) for group in blocked),
    }


def validate(stats: dict[str, int]) -> None:
    wb = load_workbook(OUTPUT, read_only=True, data_only=False)
    expected_sheets = [
        "IMPORT_READY_78",
        "REVIEW_54_TESTS",
        "REVIEW_329_RULES",
        "LAB_TEST_MAP_132",
        "UNIT_MAP_21",
        "README",
    ]
    if wb.sheetnames != expected_sheets:
        raise AssertionError(f"Unexpected sheets: {wb.sheetnames}")
    ws = wb["IMPORT_READY_78"]
    rows = ws.iter_rows(values_only=True)
    headers = list(next(rows))
    if headers != IMPORT_HEADERS:
        raise AssertionError(f"Unexpected import headers: {headers}")

    seen_codes: set[str] = set()
    rule_count = 0
    for row in rows:
        record = dict(zip(headers, row))
        service = json.loads(record["service_type"])
        section = json.loads(record["section"])
        lab_item = json.loads(record["lab_item"])
        ranges = json.loads(record["reference_ranges"])
        if service.get("value") != SERVICE_TYPE["value"]:
            raise AssertionError("service_type value mismatch")
        if section.get("value") != SECTION["value"] or section.get("code") != "BC":
            raise AssertionError("section value mismatch")
        code = str(record["source_test_code"])
        if code in seen_codes:
            raise AssertionError(f"Duplicate parent Test code {code}")
        seen_codes.add(code)
        if lab_item.get("value") != code:
            raise AssertionError(f"lab_item mismatch for Test {code}")
        if not isinstance(ranges, list) or not ranges:
            raise AssertionError(f"reference_ranges must be a non-empty array for Test {code}")
        for rule in ranges:
            if rule.get("physiological_type") not in PHYSIOLOGICAL_TYPE.values():
                raise AssertionError(f"Invalid physiological_type for Test {code}")
            if rule.get("initial_age_unit") not in INITIAL_AGE_UNIT.values():
                raise AssertionError(f"Invalid initial_age_unit for Test {code}")
            if rule.get("final_age_unit") not in FINAL_AGE_UNIT.values():
                raise AssertionError(f"Invalid final_age_unit for Test {code}")
            if rule.get("numeric_range_operator") not in NUMERIC_OPERATOR.values():
                raise AssertionError(f"Invalid numeric operator for Test {code}")
            if rule.get("range_type") not in RANGE_TYPE.values():
                raise AssertionError(f"Invalid range type for Test {code}")
            if rule.get("first_value") is None:
                raise AssertionError(f"Missing first_value for Test {code}")
            unit = rule.get("range_unit_id")
            if unit is not None and not (unit.get("value") and unit.get("label")):
                raise AssertionError(f"Invalid range_unit_id for Test {code}")
        rule_count += len(ranges)

    if len(seen_codes) != stats["ready_tests"] or rule_count != stats["ready_rules"]:
        raise AssertionError("Ready row/rule totals do not match build statistics")
    if wb["REVIEW_54_TESTS"].max_row - 1 != stats["blocked_tests"]:
        raise AssertionError("Blocked Test total mismatch")
    if wb["REVIEW_329_RULES"].max_row - 1 != stats["blocked_rules"]:
        raise AssertionError("Blocked rule total mismatch")


def main() -> None:
    stats = build()
    validate(stats)
    print(compact_json({"output": str(OUTPUT), **stats, "validation": "PASS"}))


if __name__ == "__main__":
    main()
