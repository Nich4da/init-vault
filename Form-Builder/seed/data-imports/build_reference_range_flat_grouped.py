#!/usr/bin/env python3
"""Save the user-confirmed reference form revision and build its flat import sheet.

The form revision is derived from the previous exported form so every widget keeps
its complete options object.  The workbook uses the initCraft flat Sub Form shape:
one parent row starts each Test group and continuation rows hold additional
``reference_ranges.*`` children.
"""

from __future__ import annotations

import json
from copy import deepcopy
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, Side

import build_reference_range_import as source_builder


WORKSPACE = Path("/Users/nichada/Documents/codex-backup")
CURRENT_FORM = Path("/Users/nichada/Documents/init-vault/HIS/sdform_module/reference.json")
LATEST_FORM_COPY = WORKSPACE / "reference_latest_2026-08-26.json"
READY_SOURCE = WORKSPACE / "Reference_Range_Chem_initCraft_IMPORT_ONE_SHEET_2026-08-26.xlsx"
OUTPUT = WORKSPACE / "Reference_Range_Chem_initCraft_IMPORT_FLAT_GROUPED_2026-08-26.xlsx"

ROOT_HEADERS = [
    "source_test_code",
    "section",
    "source_test_reference",
    "lab_item",
    "method",
    "revision_no",
    "is_active",
]
RANGE_NAMES = [
    "priority",
    "physiological_type",
    "range_unit_id",
    "numeric_range_operator",
    "initial_age",
    "initial_age_unit",
    "final_age",
    "final_age_unit",
    "range_type",
    "first_value",
    "second_value",
    "min_result_status_id",
    "max_result_status_id",
]
RANGE_HEADERS = [f"reference_ranges.{name}" for name in RANGE_NAMES]
TAIL_HEADERS = ["change_reason", "notes"]
HEADERS = ROOT_HEADERS + RANGE_HEADERS + TAIL_HEADERS

SECTION_ON_CHANGE = """// try {
//   const formRef = this.getFormRef();
//   const sectionCode = !!value && !!value.code ? value.code : '';
//   const sectionId = !!value && !!value.value ? value.value : '';
//
//   // :xq prefix = addParams แทรกค่าดิบ และตัดทิ้งทั้งท่อนเมื่อไม่มีค่า (SqlParser.ts:389,414)
//   // → ไม่เลือก section = ไม่มีเงื่อนไข section ใน where เลย = แสดงทั้งหมด
//   formRef.customParams.xqLabSection = !!sectionCode ? "AND `section.value` = '" + sectionCode + "'" : '';
//   formRef.customParams.xqSubOrderSection = !!sectionId ? "AND `section.value` = CONVERT('" + sectionId + "', 'objectId')" : '';
//
//   // ปิด dependEnable = ไม่มีใครล้าง cache ให้ → dropdown จะค้างรายการของ section เก่า
//   // (handleRemote จะ return ทันทีถ้า itemsList ยังมีของ — SdSelectRemoteList.vue:621)
//   const clearList = (fieldRef) => {
//     if (!!fieldRef && !!fieldRef.fieldEditor && !!fieldRef.fieldEditor.selectInput) {
//       fieldRef.fieldEditor.selectInput.handleClearList();
//     }
//   };
//
//   clearList(formRef.getFieldRef('sub_order'));
//   formRef.findFieldNameInSubForm('lab_item').forEach((refName) => clearList(formRef.getFieldRef(refName)));
// } catch (error) {
//   console.error('section filter sync failed', error);
// }"""

CSS_CODE = """  /* Sub Form ใช้สีและหน้าตาของระบบตามปกติ */
  .rr-subform {
    margin: 0 !important;
    padding: 0 !important;
    background: transparent !important;
    border: 0 !important;
    border-radius: 0 !important;
    box-shadow: none !important;
  }

  /* ซ่อนหัวคอลัมน์อัตโนมัติ แต่เก็บปุ่ม + ไว้ */
  .rr-subform .header-row {
    justify-content: flex-end !important;
    margin: 0 0 8px !important;
    padding: 0 !important;
  }

  .rr-subform .header-row .fields-header-column {
    display: none !important;
  }

  .rr-subform .header-row .action-header-column {
    flex: 0 0 auto !important;
    width: auto !important;
    max-width: none !important;
    margin-left: auto !important;
    padding: 0 5px !important;
  }

  /* แต่ละรายการ Reference Range */
  .rr-subform .sub-form-row {
    align-items: flex-start !important;
    margin: 0 !important;
    padding: 12px 0 20px !important;
    background: transparent !important;
    border: 0 !important;
    border-radius: 0 !important;
    box-shadow: none !important;
  }

  /* เว้นระหว่างรายการชุดที่ 1, 2, 3 */
  .rr-subform .sub-form-row + .sub-form-row {
    margin-top: 18px !important;
    padding-top: 18px !important;
  }

  /* ระยะห่างของช่องภายในรายการ */
  .rr-subform .sub-form-row .el-row {
    align-items: flex-start !important;
    row-gap: 14px !important;
  }

  .rr-subform .sub-form-field-column {
    margin-bottom: 4px !important;
    padding-left: 8px !important;
    padding-right: 8px !important;
  }

  /* แสดง label ภายในทุกรายการ */
  .rr-subform
  .sub-form-field-column.hide-label
  .el-form-item__label {
    display: flex !important;
    visibility: visible !important;
    opacity: 1 !important;
    align-items: center !important;
    justify-content: flex-start !important;
    width: 100% !important;
    height: auto !important;
    min-height: 20px !important;
    margin: 0 0 6px !important;
    padding: 0 !important;
    overflow: visible !important;
    text-align: left !important;
    line-height: 20px !important;
  }

  /* บังคับ label ชิดซ้าย */
  .rr-subform .el-form-item__label {
    justify-content: flex-start !important;
    text-align: left !important;
  }

  .rr-subform .el-form-item__content {
    width: 100% !important;
    margin-left: 0 !important;
  }

  /* ใช้รูปลักษณ์ input/select จากระบบเดิม */
  .rr-subform .el-input,
  .rr-subform .el-input-number,
  .rr-subform .el-select {
    width: 100% !important;
  }

  /* ซ่อนพื้นที่คอลัมน์ของ field ทางเทคนิค */
  .rr-subform
  .sub-form-field-column:has(.rr-hide-col) {
    display: none !important;
  }

  /* ซ่อนเฉพาะหัว # แต่คง #1, #2 ของแต่ละรายการ */
  .rr-subform
  .header-row
  .action-header-column.action-label {
    display: none !important;
  }

  /* Tablet */
  @media (max-width: 900px) {
    .rr-subform .sub-form-field-column {
      flex: 0 0 50% !important;
      width: 50% !important;
      max-width: 50% !important;
    }
  }

  /* Mobile */
  @media (max-width: 600px) {
    .rr-subform .sub-form-field-column {
      flex: 0 0 100% !important;
      width: 100% !important;
      max-width: 100% !important;
    }
  }"""


def walk(value: Any):
    if isinstance(value, dict):
        yield value
        for child in value.values():
            yield from walk(child)
    elif isinstance(value, list):
        for child in value:
            yield from walk(child)


def widget(document: dict[str, Any], name: str) -> dict[str, Any]:
    matches = [node for node in walk(document) if node.get("options", {}).get("name") == name]
    if len(matches) != 1:
        raise AssertionError(f"Expected one widget named {name!r}, found {len(matches)}")
    return matches[0]


def make_latest_form() -> dict[str, Any]:
    document = json.loads(CURRENT_FORM.read_text(encoding="utf-8"))
    latest = deepcopy(document)

    # Reuse complete exported widget objects and only rearrange them as confirmed.
    section = widget(latest, "section")
    lab_item = widget(latest, "lab_item")
    source_code = widget(latest, "source_test_code")
    source_reference = widget(latest, "source_test_reference")
    revision = widget(latest, "revision_no")
    method = widget(latest, "method")
    is_active = widget(latest, "is_active")

    details = widget(latest, "reference_details")
    grid1, grid2, grid3 = details["fields"]
    grid1["cols"][0]["fields"] = [section, lab_item]
    grid1["cols"][1]["fields"] = [source_code, source_reference]
    grid2["cols"][0]["fields"] = [revision, method]
    grid2["cols"][1]["fields"] = [is_active]
    grid3["cols"][0]["fields"] = []
    grid3["cols"][1]["fields"] = []

    section["options"].update(
        where="",
        dependEnable=False,
        dependField="service_type",
        onChange=SECTION_ON_CHANGE,
    )
    method["options"]["placeholder"] = ""

    option_updates = {
        "priority": {"columnSpan": 4},
        "physiological_type": {"columnSpan": 7},
        "range_unit_id": {"columnSpan": 6, "placeholder": "หน่วย"},
        "numeric_range_operator": {
            "columnSpan": 7,
            "optionItems": [
                {"label": "Inclusive limits", "value": "Inclusive limits"},
                {"value": "Exclusive limits", "label": "Exclusive limits"},
                {"label": "<", "value": "<"},
                {"value": "<=", "label": "<="},
                {"value": ">", "label": ">"},
                {"label": ">=", "value": ">="},
            ],
        },
        "initial_age": {"columnSpan": 5, "placeholder": "อายุแรกเริ่ม"},
        "initial_age_unit": {
            "columnSpan": 7,
            "placeholder": "หน่วยอายุแรกเริ่ม",
            "optionItems": [
                {"label": "Days", "value": "DAYS"},
                {"label": "Months", "value": "MONTHS"},
                {"label": "Years", "value": "YEARS"},
            ],
        },
        "final_age": {"columnSpan": 5, "placeholder": "อายุสุดท้าย"},
        "final_age_unit": {
            "columnSpan": 7,
            "placeholder": "หน่วยอายุสุดท้าย",
            "optionItems": [
                {"label": "Days", "value": "DAYS"},
                {"label": "Months", "value": "MONTHS"},
                {"label": "Years", "value": "YEARS"},
            ],
        },
        "range_type": {
            "columnSpan": 8,
            "optionItems": [
                {"label": "NormalNumeric", "value": "NormalNumeric"},
                {"label": "CriticalNumeric", "value": "CriticalNumeric"},
            ],
        },
        "first_value": {"columnSpan": 8},
        "second_value": {"columnSpan": 8},
        "min_result_status_id": {"columnSpan": 6, "placeholder": "min result status id"},
        "max_result_status_id": {"columnSpan": 6, "placeholder": "max result status id"},
    }
    for name, updates in option_updates.items():
        widget(latest, name)["options"].update(updates)

    subform = widget(latest, "reference_ranges")
    subform["options"].update(
        showRowNumber=True,
        customClass=["rr-subform", "sub-form-field-column:has("],
    )
    latest["formConfig"]["cssCode"] = CSS_CODE

    if any(node.get("options", {}).get("name") == "service_type" for node in walk(latest)):
        raise AssertionError("service_type still exists after latest-form transformation")
    LATEST_FORM_COPY.write_text(
        json.dumps(latest, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )
    return latest


def compact_json(value: Any) -> str:
    return json.dumps(value, ensure_ascii=False, separators=(",", ":"))


def source_groups_by_code() -> dict[str, dict[str, Any]]:
    groups = source_builder.load_source()
    result = {group["code"]: group for group in groups}
    if len(result) != len(groups):
        raise AssertionError("Duplicate source Test code")
    return result


def load_ready_records() -> list[dict[str, Any]]:
    workbook = load_workbook(READY_SOURCE, read_only=True, data_only=True)
    if workbook.sheetnames != ["IMPORT_READY"]:
        raise AssertionError("The audited ready source must contain one IMPORT_READY sheet")
    rows = workbook["IMPORT_READY"].iter_rows(values_only=True)
    headers = list(next(rows))
    return [dict(zip(headers, row)) for row in rows]


def build_flat_workbook() -> tuple[int, int]:
    groups = source_groups_by_code()
    records = load_ready_records()
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "IMPORT_READY"
    sheet.append(HEADERS)

    seen_codes: set[str] = set()
    total_rules = 0
    group_start_rows: list[int] = []
    for record in records:
        code = str(record["source_test_code"])
        if code in seen_codes:
            raise AssertionError(f"Duplicate ready Test code: {code}")
        seen_codes.add(code)
        source_group = groups[code]
        old_rules = json.loads(record["reference_ranges"])
        if len(old_rules) != len(source_group["rules"]):
            raise AssertionError(f"Rule count mismatch for Test {code}")
        section = json.loads(record["section"])
        lab_item = json.loads(record["lab_item"])
        group_start_rows.append(sheet.max_row + 1)

        for index, rule in enumerate(source_group["rules"]):
            first = index == 0
            row = {
                "source_test_code": code if first else None,
                "section": compact_json(section) if first else None,
                "source_test_reference": source_group["reference"] if first else None,
                "lab_item": compact_json(lab_item) if first else None,
                "method": source_group["method"] if first else None,
                "revision_no": 1 if first else None,
                "is_active": True if first else None,
                "reference_ranges.priority": rule["priority"],
                "reference_ranges.physiological_type": {
                    "ALL": "All",
                    "M": "M",
                    "F": "F",
                }[rule["physiological_type_source"]],
                "reference_ranges.range_unit_id": rule["range_unit_source"],
                "reference_ranges.numeric_range_operator": rule["numeric_operator_source"],
                "reference_ranges.initial_age": rule["initial_age"],
                "reference_ranges.initial_age_unit": rule["initial_age_unit_source"],
                "reference_ranges.final_age": rule["final_age"],
                "reference_ranges.final_age_unit": rule["final_age_unit_source"],
                "reference_ranges.range_type": rule["range_type_source"],
                "reference_ranges.first_value": rule["first_value"],
                "reference_ranges.second_value": rule["second_value"],
                "reference_ranges.min_result_status_id": rule["min_result_status_id"],
                "reference_ranges.max_result_status_id": rule["max_result_status_id"],
                "change_reason": None,
                "notes": None,
            }
            sheet.append([row[header] for header in HEADERS])
            total_rules += 1

    header_font = Font(bold=True)
    group_border = Border(top=Side(style="medium"))
    for cell in sheet[1]:
        cell.font = header_font
        cell.alignment = Alignment(vertical="top", wrap_text=True)
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for row_number in group_start_rows:
        for cell in sheet[row_number]:
            cell.border = group_border

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for column in range(1, len(HEADERS) + 1):
        letter = sheet.cell(1, column).column_letter
        sheet.column_dimensions[letter].width = 22
    sheet.column_dimensions["B"].width = 55
    sheet.column_dimensions["D"].width = 95
    workbook.save(OUTPUT)
    return len(records), total_rules


def validate_workbook(expected_groups: int, expected_rules: int) -> None:
    workbook = load_workbook(OUTPUT, read_only=False, data_only=False)
    if workbook.sheetnames != ["IMPORT_READY"]:
        raise AssertionError(f"Expected exactly one sheet, got {workbook.sheetnames}")
    sheet = workbook["IMPORT_READY"]
    if sheet.max_row != expected_rules + 1 or sheet.max_column != len(HEADERS):
        raise AssertionError("Workbook dimensions do not match expected flat rows")
    if sheet.merged_cells.ranges:
        raise AssertionError("Import workbook must not contain merged cells")
    rows = sheet.iter_rows(values_only=True)
    if list(next(rows)) != HEADERS:
        raise AssertionError("Workbook headers differ from the latest form paths")

    group_count = 0
    current: dict[str, Any] | None = None
    reconstructed: list[dict[str, Any]] = []
    for values in rows:
        row = dict(zip(HEADERS, values))
        if any(isinstance(value, str) and value.startswith("=") for value in values if value is not None):
            raise AssertionError("Import workbook contains a formula")
        if row["source_test_code"] is not None:
            group_count += 1
            current = {
                "source_test_code": str(row["source_test_code"]),
                "section": json.loads(row["section"]),
                "lab_item": json.loads(row["lab_item"]),
                "reference_ranges": [],
            }
            reconstructed.append(current)
        if current is None:
            raise AssertionError("Continuation row appears before the first parent row")
        rule = {name: row[f"reference_ranges.{name}"] for name in RANGE_NAMES}
        required = [
            "priority",
            "physiological_type",
            "numeric_range_operator",
            "initial_age_unit",
            "final_age_unit",
            "range_type",
            "first_value",
        ]
        if any(rule[name] in (None, "") for name in required):
            raise AssertionError(f"Required Sub Form value missing: {rule}")
        if rule["physiological_type"] not in {"All", "M", "F"}:
            raise AssertionError("Invalid physiological_type")
        if rule["initial_age_unit"] not in {"DAYS", "MONTHS", "YEARS"}:
            raise AssertionError("Invalid initial_age_unit")
        if rule["final_age_unit"] not in {"DAYS", "MONTHS", "YEARS"}:
            raise AssertionError("Invalid final_age_unit")
        if rule["numeric_range_operator"] not in {
            "Inclusive limits", "Exclusive limits", "<", "<=", ">", ">="
        }:
            raise AssertionError("Invalid numeric_range_operator")
        if rule["range_type"] not in {"NormalNumeric", "CriticalNumeric"}:
            raise AssertionError("Invalid range_type")
        current["reference_ranges"].append(rule)

    if group_count != expected_groups or len(reconstructed) != expected_groups:
        raise AssertionError("Parent group count mismatch")
    if sum(len(item["reference_ranges"]) for item in reconstructed) != expected_rules:
        raise AssertionError("Reconstructed Sub Form row count mismatch")


def main() -> None:
    latest = make_latest_form()
    groups, rules = build_flat_workbook()
    validate_workbook(groups, rules)
    names = {
        node.get("options", {}).get("name")
        for node in walk(latest)
        if node.get("options", {}).get("name")
    }
    print(f"latest_form={LATEST_FORM_COPY}")
    print(f"service_type_present={'service_type' in names}")
    print(f"workbook={OUTPUT}")
    print(f"sheets=1 groups={groups} subform_rows={rules} columns={len(HEADERS)}")


if __name__ == "__main__":
    main()
