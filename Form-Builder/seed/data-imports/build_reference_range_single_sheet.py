#!/usr/bin/env python3
"""Create a one-sheet initCraft import workbook for CHE reference ranges.

The previous workbook intentionally included audit/review worksheets. initCraft's
record import needs a workbook containing only the actual import table, so this
script validates the current Numbers source against that audited workbook and
then emits just the ready parent records.
"""

from __future__ import annotations

import json
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font

import build_reference_range_import as source_builder


WORKSPACE = Path("/Users/nichada/Documents/codex-backup")
AUDITED_WORKBOOK = WORKSPACE / "Reference_Range_Chem_initCraft_Import_2026-08-26.xlsx"
FORM_JSON = Path("/Users/nichada/Documents/init-vault/HIS/sdform_module/reference.json")
OUTPUT = WORKSPACE / "Reference_Range_Chem_initCraft_IMPORT_ONE_SHEET_2026-08-26.xlsx"

IMPORT_SHEET = "IMPORT_READY_78"
EXPECTED_HEADERS = [
    "service_type",
    "section",
    "source_test_code",
    "source_test_reference",
    "lab_item",
    "revision_no",
    "is_active",
    "reference_ranges",
]
EXPECTED_RANGE_FIELDS = [
    "priority",
    "physiological_type",
    "range_unit_id",
    "numeric_range_operator",
    "initial_age",
    "initial_age_unit",
    "final_age",
    "final_age_unit",
    "range_type",
    "first_value",
    "second_value",
    "min_result_status_id",
    "max_result_status_id",
]


def normalize(value: Any) -> Any:
    """Normalize spreadsheet scalar values for stable comparisons."""
    if value is None:
        return None
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return source_builder.number_text(value)
    text = str(value).strip()
    return text or None


def iter_form_objects(value: Any):
    if isinstance(value, dict):
        yield value
        for child in value.values():
            yield from iter_form_objects(child)
    elif isinstance(value, list):
        for child in value:
            yield from iter_form_objects(child)


def validate_form_pattern() -> None:
    document = json.loads(FORM_JSON.read_text(encoding="utf-8"))
    option_names = {
        item["options"]["name"]
        for item in iter_form_objects(document)
        if isinstance(item.get("options"), dict) and item["options"].get("name")
    }
    missing_headers = [header for header in EXPECTED_HEADERS if header not in option_names]
    if missing_headers:
        raise AssertionError(f"Import headers absent from reference.json: {missing_headers}")

    subform = next(
        (
            item
            for item in iter_form_objects(document)
            if item.get("component") == "sub-form"
            and item.get("options", {}).get("name") == "reference_ranges"
        ),
        None,
    )
    if subform is None:
        raise AssertionError("reference_ranges Sub Form not found in reference.json")
    actual_range_fields = [field.get("options", {}).get("name") for field in subform.get("fields", [])]
    if actual_range_fields != EXPECTED_RANGE_FIELDS:
        raise AssertionError(
            "reference_ranges field pattern changed: "
            f"expected={EXPECTED_RANGE_FIELDS}, actual={actual_range_fields}"
        )


def audited_rows(workbook) -> tuple[list[str], list[dict[str, Any]]]:
    if IMPORT_SHEET not in workbook.sheetnames:
        raise AssertionError(f"Missing audited import sheet: {IMPORT_SHEET}")
    worksheet = workbook[IMPORT_SHEET]
    rows = worksheet.iter_rows(values_only=True)
    headers = list(next(rows))
    if headers != EXPECTED_HEADERS:
        raise AssertionError(f"Unexpected audited headers: {headers}")
    return headers, [dict(zip(headers, row)) for row in rows]


def compare_ready_record(record: dict[str, Any], source_group: dict[str, Any]) -> int:
    code = str(record["source_test_code"])
    if code != source_group["code"]:
        raise AssertionError(f"Source code mismatch: {code}")
    if normalize(record["source_test_reference"]) != normalize(source_group["reference"]):
        raise AssertionError(f"Source reference changed for Test {code}")

    service = json.loads(record["service_type"])
    section = json.loads(record["section"])
    lab_item = json.loads(record["lab_item"])
    ranges = json.loads(record["reference_ranges"])

    if service.get("value") != source_builder.SERVICE_TYPE["value"]:
        raise AssertionError(f"service_type mismatch for Test {code}")
    if section.get("value") != source_builder.SECTION["value"] or section.get("code") != "BC":
        raise AssertionError(f"section mismatch for Test {code}")
    if str(lab_item.get("value")) != code:
        raise AssertionError(f"lab_item mismatch for Test {code}")
    if list(ranges[0].keys()) != EXPECTED_RANGE_FIELDS:
        raise AssertionError(f"Sub Form key order/pattern mismatch for Test {code}")
    if len(ranges) != len(source_group["rules"]):
        raise AssertionError(f"Rule count changed for Test {code}")

    for imported, source in zip(ranges, source_group["rules"]):
        expected = {
            "priority": source["priority"],
            "physiological_type": source_builder.PHYSIOLOGICAL_TYPE[source["physiological_type_source"]],
            "numeric_range_operator": source_builder.NUMERIC_OPERATOR[source["numeric_operator_source"]],
            "initial_age": source["initial_age"],
            "initial_age_unit": source_builder.INITIAL_AGE_UNIT[source["initial_age_unit_source"]],
            "final_age": source["final_age"],
            "final_age_unit": source_builder.FINAL_AGE_UNIT[source["final_age_unit_source"]],
            "range_type": source_builder.RANGE_TYPE[source["range_type_source"]],
            "first_value": source["first_value"],
            "second_value": source["second_value"],
            "min_result_status_id": source["min_result_status_id"],
            "max_result_status_id": source["max_result_status_id"],
        }
        for field_name, expected_value in expected.items():
            if normalize(imported.get(field_name)) != normalize(expected_value):
                raise AssertionError(
                    f"Current Numbers source differs at Test {code}, source row "
                    f"{source['source_row']}, field {field_name}"
                )

        source_unit = source["range_unit_source"]
        imported_unit = imported.get("range_unit_id")
        if source_unit is None:
            if imported_unit is not None:
                raise AssertionError(f"Unexpected unit for Test {code}, source row {source['source_row']}")
        elif not isinstance(imported_unit, dict) or normalize(imported_unit.get("label")) != normalize(source_unit):
            raise AssertionError(f"Unit changed for Test {code}, source row {source['source_row']}")
        elif not imported_unit.get("value"):
            raise AssertionError(f"Unit master id missing for Test {code}, source row {source['source_row']}")

    return len(ranges)


def validate_current_numbers(workbook, records: list[dict[str, Any]]) -> tuple[int, int, int, int]:
    groups = source_builder.load_source()
    groups_by_code = {group["code"]: group for group in groups}
    if len(groups_by_code) != len(groups):
        raise AssertionError("Duplicate Test group code in current Numbers source")

    ready_codes = {str(record["source_test_code"]) for record in records}
    if len(ready_codes) != len(records):
        raise AssertionError("Duplicate parent Test code in audited import sheet")

    review_ws = workbook["REVIEW_54_TESTS"]
    review_rows = review_ws.iter_rows(values_only=True)
    review_headers = list(next(review_rows))
    blocked = [dict(zip(review_headers, row)) for row in review_rows]
    blocked_codes = {str(row["source_test_code"]) for row in blocked}

    if ready_codes & blocked_codes:
        raise AssertionError("A Test code appears in both ready and review groups")
    if set(groups_by_code) != ready_codes | blocked_codes:
        raise AssertionError("Current Numbers Test groups differ from the audited workbook")

    for row in blocked:
        code = str(row["source_test_code"])
        group = groups_by_code[code]
        if normalize(row["source_test_reference"]) != normalize(group["reference"]):
            raise AssertionError(f"Blocked Test reference changed for {code}")
        if int(row["rule_count"]) != len(group["rules"]):
            raise AssertionError(f"Blocked Test rule count changed for {code}")

    ready_rule_count = 0
    for record in records:
        code = str(record["source_test_code"])
        ready_rule_count += compare_ready_record(record, groups_by_code[code])

    source_rule_count = sum(len(group["rules"]) for group in groups)
    blocked_rule_count = sum(len(groups_by_code[code]["rules"]) for code in blocked_codes)
    if source_rule_count != ready_rule_count + blocked_rule_count:
        raise AssertionError("Source/ready/blocked rule totals do not reconcile")

    return len(groups), source_rule_count, len(records), ready_rule_count


def write_one_sheet(headers: list[str], records: list[dict[str, Any]]) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "IMPORT_READY"
    worksheet.append(headers)
    for record in records:
        worksheet.append([record[header] for header in headers])

    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(vertical="top", wrap_text=True)
    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    widths = [45, 55, 20, 28, 95, 12, 12, 120]
    for index, width in enumerate(widths, start=1):
        worksheet.column_dimensions[worksheet.cell(row=1, column=index).column_letter].width = width
    workbook.save(OUTPUT)


def validate_output(expected_rows: int, expected_rules: int) -> None:
    workbook = load_workbook(OUTPUT, read_only=False, data_only=False)
    if workbook.sheetnames != ["IMPORT_READY"]:
        raise AssertionError(f"Output must contain exactly one sheet: {workbook.sheetnames}")
    worksheet = workbook["IMPORT_READY"]
    if worksheet.max_row != expected_rows + 1 or worksheet.max_column != len(EXPECTED_HEADERS):
        raise AssertionError("Output dimensions do not match the validated import data")
    if worksheet.merged_cells.ranges:
        raise AssertionError("Import sheet must not contain merged cells")

    rows = worksheet.iter_rows(values_only=True)
    headers = list(next(rows))
    if headers != EXPECTED_HEADERS:
        raise AssertionError(f"Output headers changed: {headers}")

    seen_codes: set[str] = set()
    rule_count = 0
    for row in rows:
        record = dict(zip(headers, row))
        if any(isinstance(value, str) and value.startswith("=") for value in row if value is not None):
            raise AssertionError("Import sheet must not contain formulas")
        code = str(record["source_test_code"])
        if code in seen_codes:
            raise AssertionError(f"Duplicate parent record for Test {code}")
        seen_codes.add(code)
        json.loads(record["service_type"])
        json.loads(record["section"])
        json.loads(record["lab_item"])
        ranges = json.loads(record["reference_ranges"])
        if not isinstance(ranges, list) or not ranges:
            raise AssertionError(f"reference_ranges is not a non-empty array for Test {code}")
        for rule in ranges:
            if list(rule.keys()) != EXPECTED_RANGE_FIELDS:
                raise AssertionError(f"Sub Form field pattern mismatch for Test {code}")
        rule_count += len(ranges)

    if len(seen_codes) != expected_rows or rule_count != expected_rules:
        raise AssertionError("Output parent/rule totals do not match validation totals")


def main() -> None:
    validate_form_pattern()
    audited = load_workbook(AUDITED_WORKBOOK, read_only=True, data_only=False)
    headers, records = audited_rows(audited)
    source_tests, source_rules, ready_tests, ready_rules = validate_current_numbers(audited, records)
    write_one_sheet(headers, records)
    validate_output(ready_tests, ready_rules)
    print(
        json.dumps(
            {
                "output": str(OUTPUT),
                "sheets": 1,
                "source_tests": source_tests,
                "source_rules": source_rules,
                "import_parent_records": ready_tests,
                "import_reference_range_rules": ready_rules,
                "validation": "PASS",
            },
            ensure_ascii=False,
            separators=(",", ":"),
        )
    )


if __name__ == "__main__":
    main()
